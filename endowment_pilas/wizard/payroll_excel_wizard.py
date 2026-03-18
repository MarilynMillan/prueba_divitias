from odoo import fields, models, _
from odoo.exceptions import UserError
import io
import os
import base64
from datetime import datetime, date
from odoo.modules import get_module_resource



class PayrollExcelWizard(models.TransientModel):
    _name = 'hr.payroll.excel.wizard'
    _description = 'Wizard para Reporte de Nómina a Excel'

    date_from = fields.Date(string="Fecha Desde")
    date_to = fields.Date(string="Fecha Hasta")
    payslip_run_id = fields.Many2one('hr.payslip.run', string="Lote de Nómina")

    plantilla_excel = fields.Binary(string='Plantilla Excel', help='Sube la plantilla de Excel para usar en el reporte. Si está vacío, se usará la plantilla por defecto.')
    plantilla_excel_name = fields.Char(string='Nombre Archivo Plantilla')

    def action_generate_excel_report(self):
        """
        Genera el reporte Excel de recibos de nómina filtrados.
        """
        domain = []

        if self.date_from:
            domain.append(('date_from', '>=', self.date_from))
        if self.date_to:
            domain.append(('date_to', '<=', self.date_to))


        if self.payslip_run_id:
            domain.append(('payslip_run_id', '=', self.payslip_run_id.id))

        # Si no hay filtros, el dominio estará vacío y traerá todos los recibos.
        # Puedes añadir una validación aquí si quieres forzar al menos un filtro.
        if not domain:
            raise UserError(_("Por favor, selecciona al menos un filtro de fecha o un lote de nómina."))

        payslips = self.env['hr.payslip'].search(domain)

        if not payslips:
            raise UserError(_("No se encontraron recibos de nómina para los filtros seleccionados."))

        import io
        import base64
        import os
        import openpyxl
        from openpyxl.styles import Alignment
        from odoo.modules import get_module_resource

        output = io.BytesIO()

        # 1. CARGAR LA PLANTILLA (El molde que ya tiene todo el diseño)
        if self.plantilla_excel:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.plantilla_excel)))
                sheet = wb.active
            except Exception as e:
                raise UserError(f"Error al abrir la plantilla subida: {str(e)}")
        else:
            path = get_module_resource('endowment_pilas', 'data', 'DIVITIASSAS_1.xlsx')
            if not path or not os.path.exists(path):
                 raise UserError(f"No se encontró la plantilla en: {path}")

            try:
                wb = openpyxl.load_workbook(path)
                sheet = wb.active
            except Exception as e:
                raise UserError(f"Error al abrir la plantilla: {str(e)}")

        # 2. DEFINIR SOLO EL ESTILO DE DATOS
        left_alignment = Alignment(horizontal='left', vertical='center')

        

        # 3. LLENADO DE DATOS (Empezamos en la fila 19 porque la 17 y 18 son encabezados)
        row_num = 19
        row_counter = 1

        for payslip in payslips:
            employee = payslip.employee_id
            contract = payslip.contract_id
            correction_status_value = payslip.correction_status if hasattr(payslip, 'correction_status') else 'No'


            ############################## type identification#########################
            # Obtenemos el registro del tipo de identificación

            mapeo_id = {
                'Cédula de ciudadanía': 'CC',
                'Cédula de extranjería': 'CE',
                'Tarjeta de Identidad': 'TI',
                'Registro Civil': 'RC',
                'Pasaporte': 'PA',
                'Permiso por Protección Temporal': 'PT', # El que viste en la lista
                'PEP (Permiso Especial de Permanencia)': 'PE',
                'NIT': 'NI',
                'ID Extranjera': 'CE',
                'Documento de identificación extranjero': 'CD'
            }

            # 1. Obtenemos el registro
            tipo_doc_rec = employee.employee_address_home.l10n_latam_identification_type_id

            # 2. Extraemos el nombre (string)
            nombre_largo = tipo_doc_rec.name or ''

            # 3. Aplicamos el mapeo que definimos antes para que salga "CC", "CE", etc.
            # mapeo_id es el diccionario que definimos en el paso anterior
            tipo_doc_abreviado = mapeo_id.get(nombre_largo, nombre_largo)
            #####################################################################
            tipo_cotizante_excel_label = '' 
            if contract:
                if contract.pila_tipo_trabajador_id:
                    # Usamos el nombre del nuevo campo configurable, si tiene algo (como '1.Dependiente')
                    tipo_cotizante_excel_label = contract.pila_tipo_trabajador_id.name
                elif contract.tipo_trabajador:
                    # --- CORRECCIÓN AQUÍ ---
                    # Método recomendado para obtener la etiqueta legible del campo Selection
                    tipo_cotizante_excel_label = dict(contract._fields['tipo_trabajador'].selection).get(contract.tipo_trabajador, '')

            sub_cotizante_excel_label = '' 
            if contract:
                if contract.pila_subtipo_trabajador_id:
                    sub_cotizante_excel_label = contract.pila_subtipo_trabajador_id.name
                elif contract.sub_tipo_trabajador:
                    # --- CORRECCIÓN AQUÍ ---
                    # Método recomendado para obtener la etiqueta legible del campo Selection
                    sub_cotizante_excel_label = dict(contract._fields['sub_tipo_trabajador'].selection).get(contract.sub_tipo_trabajador, '')

            # --- NUEVO: Lógica para 'Horas Laboradas' ---

            dias_cotizados_pension = 0.0
            dias_cotizados_salud = 0.0
            dias_cotizados_arl = 0.0
            dias_cotizados_ccf = 0.0
            
            horas_laboradas = 0.0

            # Iteramos sobre todas las líneas de días trabajados/novedades
            for worked_day_line in payslip.worked_days_line_ids:
                days = worked_day_line.number_of_days
                code = worked_day_line.code
                
                # 1. ACUMULACIÓN DE HORAS LABORADAS (Solo WORK100)
                if code == 'WORK100': 
                    horas_laboradas += worked_day_line.number_of_hours

                
                # 2. ACUMULACIÓN DE DÍAS COTIZADOS 
                # Se utiliza el código de la línea del día trabajado directamente 
                # para determinar a qué columna PILA se suma.
                
                if code == 'pension':
                    dias_cotizados_pension += days
                elif code == 'salud':
                    dias_cotizados_salud += days
                elif code == 'arl':
                    dias_cotizados_arl += days
                elif code == 'ccf':
                    dias_cotizados_ccf += days


            ####################################################

            # Inicializa ambas etiquetas
            es_extranjero_label = 'No'
            es_residente_label = 'No'

            # Lógica para 'Extranjero' (Basada en País)
            if employee and employee.country_id:
                if employee.country_id.name != 'Colombia':
                    es_extranjero_label = 'Si'
                    
            if employee and employee.is_non_resident:
                es_residente_label = 'Si'

            
            
            # --- Lógica: Fecha de radicación en el exterior ---
            fecha_radicacion = ''
            if employee and getattr(employee, 'date_resident', False):
                fecha_radicacion = employee.date_resident  # Debe ser tipo Date en tu modelo

             # --- Lógica: Fecha inicio del contrato ---
            fecha_inicio_contrato = ''
            if contract and getattr(contract, 'date_start', False):
                fecha_inicio_contrato = contract.date_start  # Tipo Date normalmente


            # --- Lógica: ING (Ingreso) ---
            """valor_ing = 'NO'
            if fecha_inicio_contrato:
                # Si se definió rango desde / hasta
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_inicio_contrato <= self.date_to:
                        valor_ing = 'SI'
                # Si solo hay fecha desde
                elif self.date_from:
                    if fecha_inicio_contrato >= self.date_from:
                        valor_ing = 'SI'
                # Si solo hay fecha hasta
                elif self.date_to:
                    if fecha_inicio_contrato <= self.date_to:
                        valor_ing = 'SI'
                # Si no hay filtro de fechas, pero existe fecha de contrato
                else:
                    valor_ing = 'SI'

            
            # --- Lógica: Fecha final del contrato ---
            fecha_final_contrato = False
            if contract and getattr(contract, 'date_end', False):
                fecha_final_contrato = contract.date_end  # Tipo Date normalmente

            # --- Lógica: RET (Retiro) ---
            valor_ret = 'NO'
            if fecha_final_contrato:
                # Si se definió rango desde / hasta
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_final_contrato <= self.date_to:
                        valor_ret = 'SI'
                # Si solo hay fecha desde
                elif self.date_from:
                    if fecha_final_contrato >= self.date_from:
                        valor_ret = 'SI'
                # Si solo hay fecha hasta
                elif self.date_to:
                    if fecha_final_contrato <= self.date_to:
                        valor_ret = 'SI'
                # Si no hay filtro de fechas, pero existe fecha de contrato
                else:
                    valor_ret = 'SI'"""

            # --- Lógica: ING (Ingreso) ---
            valor_ing = 'NO'
            if fecha_inicio_contrato:
                es_periodo_ingreso = False
                # Validación de rango
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_inicio_contrato <= self.date_to:
                        es_periodo_ingreso = True
                elif self.date_from:
                    if fecha_inicio_contrato >= self.date_from:
                        es_periodo_ingreso = True
                elif self.date_to:
                    if fecha_inicio_contrato <= self.date_to:
                        es_periodo_ingreso = True
                else:
                    es_periodo_ingreso = True

                # Si está en el rango, aplicamos el concepto del contrato
                if es_periodo_ingreso:
                    # Buscamos el código del Many2one, si no hay, ponemos 'X'
                    valor_ing = contract.pila_ingreso_concepto_id.name or 'NO'

            # --- Lógica: Fecha final del contrato ---
            fecha_final_contrato = False
            if contract and getattr(contract, 'date_end', False):
                fecha_final_contrato = contract.date_end

            # --- Lógica: RET (Retiro) ---
            valor_ret = 'NO'
            if fecha_final_contrato:
                es_periodo_retiro = False
                # Validación de rango
                if self.date_from and self.date_to:
                    if self.date_from <= fecha_final_contrato <= self.date_to:
                        es_periodo_retiro = True
                elif self.date_from:
                    if fecha_final_contrato >= self.date_from:
                        es_periodo_retiro = True
                elif self.date_to:
                    if fecha_final_contrato <= self.date_to:
                        es_periodo_retiro = True
                else:
                    es_periodo_retiro = True

                # Si está en el rango, aplicamos el concepto del contrato
                if es_periodo_retiro:
                    # Buscamos el código del Many2one, si no hay, ponemos 'X'
                    valor_ret = contract.pila_retiro_concepto_id.name or 'NO'

            # Inicializar un diccionario para guardar los valores por defecto 'NO'
            # Asegúrate de que los códigos aquí coincidan con los de tu campo 'novelty_code'
            # 1. Inicializamos todas las novedades simples en 'NO' por defecto
            novelty_values = {code: 'NO' for code in ['TDE', 'TAE', 'TDP', 'TAP']} # Agrega aquí tus códigos
            
            report_start_date = self.date_from
            report_end_date = self.date_to

            afp_destino = ""
            eps_destino = ""

            # Supongamos que recorremos las administradoras configuradas en el contrato
            # o donde las tengas relacionadas (ajusta 'contract.administradora_ids' según tu modelo)
            if contract.administradoras_ids:
                for admin_line in contract.administradoras_ids:
                    # 1. ¿Es una administradora de tipo Pensión y tiene Traslado marcado?
                    if admin_line.type_entity == 'pension' and admin_line.traslado:
                        if admin_line.list_administradora_destino_id:
                            # Tomamos el nombre de la entidad destino de la lista maestra
                            afp_destino = admin_line.list_administradora_destino_id.name.upper()

                    # 2. ¿Es una administradora de tipo Salud y tiene Traslado marcado?
                    elif admin_line.type_entity == 'salud' and admin_line.traslado:
                        if admin_line.list_administradora_destino_id:
                            # Tomamos el nombre de la entidad destino de la lista maestra
                            eps_destino = admin_line.list_administradora_destino_id.name.upper()

            # Inicializar VSP con valores vacíos/por defecto
            valor_vsp = 'NO'
            fecha_vsp_str = ''

            # ----------------------------------------------------
            # 🆕 Lógica Simplificada para VSP (Variación Salario)
            # ----------------------------------------------------

            # 1. Obtenemos la fecha de cambio del campo dedicado del contrato
            fecha_cambio_sueldo = contract.date_wage_change

            if fecha_cambio_sueldo and report_start_date and report_end_date:
                
                # 2. Verificamos si la fecha de cambio cae DENTRO del rango del reporte.
                # El filtro que usa el reporte: [self.date_from, self.date_to]
                if report_start_date <= fecha_cambio_sueldo <= report_end_date:
                    
                    # Si la fecha de cambio está dentro del periodo, es VSP = 'SI'
                    valor_vsp = 'SI'
                    fecha_vsp_str = fecha_cambio_sueldo # Ya es un objeto date

            #####################################################################
            # --- Lógica: VST (Variación Transitoria de Salario) ---
            valor_vst = 'NO'
            
            # Buscamos los payslips del empleado que estén dentro del rango y realizados (state = 'done')
            # 'contract' es el objeto del contrato actual en tu bucle
            payslips = self.env['hr.payslip'].search([
                ('employee_id', '=', contract.employee_id.id),
                ('date_from', '>=', report_start_date),
                ('date_to', '<=', report_end_date),
                ('state', '=', 'done')
            ])

            if payslips:
                # Revisamos las líneas de esas nóminas (slip_ids.line_ids)
                # Buscamos si alguna línea pertenece a una regla con is_payment_transitory = True
                # Y que el monto (total) sea mayor a cero
                transitory_lines = payslips.mapped('line_ids').filtered(
                    lambda l: l.salary_rule_id.is_payment_transitory and l.total > 0
                )
                
                if transitory_lines:
                    valor_vst = 'SI'

            # ----------------------------------------------------
            # 📌 Lógica para Ausencias (Usando hr.leave - Solicitudes de Ausencia)
            # ----------------------------------------------------

            # Definición de códigos relevantes y Inicialización a 'NO'
            ABSENCE_CODES = ['SLN', 'IGE', 'LMA', 'VAL-LR', 'AVP', 'VCT', 'IRL']
            absence_novelties = {code: {'value': 'NO', 'start': '', 'end': ''} for code in ABSENCE_CODES}

            # PASO 1 y 2: Buscar Solicitudes de Ausencia (hr.leave) para el empleado.
            # --- AGREGAR: Inicialización de contadores ---
            count_lma = 0
            count_ige = 0

            # Tu búsqueda actual (sin cambios)
            leaves = self.env['hr.leave'].search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('date_from', '<=', report_end_date),
                ('date_to', '>=', report_start_date),
            ])

            if leaves:
                for leave in leaves:
                    # Tu obtención de código actual
                    novelty_code = leave.holiday_status_id.pila_novelty_code 
                    
                    if novelty_code and novelty_code in ABSENCE_CODES:
                        
                        # --- AGREGAR: Solo estas 4 líneas para el conteo ---
                        if novelty_code == 'LMA':
                            count_lma += 1
                        elif novelty_code == 'IGE':
                            count_ige += 1
                        
                        # Tu lógica de 'SI' y fechas (se mantiene igual)
                        #absence_novelties[novelty_code]['value'] = 'SI'
                        absence_novelties[novelty_code]['value'] = leave.holiday_status_id.name or 'SI'
                        
                        if novelty_code != 'AVP': 
                            start_date = leave.date_from.date()
                            end_date = leave.date_to.date()
                            absence_novelties[novelty_code]['start'] = start_date
                            absence_novelties[novelty_code]['end'] = end_date


            # Usaremos el nombre (label) para el Excel.
            # El método _get_selection_label() toma el nombre legible del campo de selección.
            correction_status_excel_value = payslip.correction_status if hasattr(payslip, 'correction_status') else 'no'

            # 2. Salario Mensual (Del campo 'wage' del contrato)
            salario_mensual = contract.wage if contract else 0.0

                    # ----------------------------------------------------
            # 📌 LÓGICA PARA SALARIO INTEGRAL Y VARIABLE
            # ----------------------------------------------------
            
            # 1. Salario Integral (True/False del contrato -> SI/NO)
            if contract and contract.wage_integral:
                integral_excel_value = 'SI'
            else:
                integral_excel_value = 'NO'

            # 2. Salario Variable (True/False del contrato -> SI/NO)
            if contract and contract.wage_variable:
                variable_excel_value = 'SI'
            else:
                variable_excel_value = 'NO'

            ##################################################################

            # ADMINISTRADORAS
            pension_admin = contract.get_admin_by_type('pension')
            salud_admin = contract.get_admin_by_type('salud')
            arl_admin = contract.get_admin_by_type('arl')
            ccf_admin = contract.get_admin_by_type('ccf')

            ##########################################################
            # Inicializamos los valores de IBC

            """ibc_values = {
                    'pension': 0.0,
                    'salud': 0.0,
                    'arl': 0.0,
                    'ccf': 0.0,
                }

            # Recorro las líneas del recibo
            for line in payslip.line_ids:
                rule = line.salary_rule_id

                if not rule:
                        continue

                # Si la regla está marcada como PENSIÓN → va a esa celda
                if rule.is_pension:
                    ibc_values['pension'] = line.total

                # Si está marcada como SALUD → va a esa celda
                if rule.is_salud:
                    ibc_values['salud'] = line.total

                # Si la regla pertenece a ARL
                if rule.is_arl:
                    ibc_values['arl'] = line.total

                # Si está marcada como CCF
                if rule.is_ccf:
                    ibc_values['ccf'] = line.total

            excel_ibc_pension = ibc_values['pension'] if ibc_values['pension'] > 0 else ""
            excel_ibc_salud   = ibc_values['salud'] if ibc_values['salud'] > 0 else ""
            excel_ibc_arl     = ibc_values['arl'] if ibc_values['arl'] > 0 else ""
            excel_ibc_ccf     = ibc_values['ccf'] if ibc_values['ccf'] > 0 else """



            # --------------------------- Lógica de Alto Riesgo ---------------------------
            alto_riesgo_selection = self.env['hr.contract'].fields_get(allfields=['alto_riesgo']
            )['alto_riesgo']['selection']

            alto_riesgo_label = ''
            if contract and contract.alto_riesgo:
                alto_riesgo_label = dict(alto_riesgo_selection).get(contract.alto_riesgo, '')

            ############################################################################
            # Reglas salariales por CODE

            """reglas_excel = {
                'valor_cotizacion_pension':       {'code': 'CotPension'},
                'valor_cotizacion_salud':         {'code': 'CotSalud'},
                'valor_cotizacion_riesgo':        {'code': 'CotRiesgo'},
                'valor_cotizacion_ccf':           {'code': 'CotCcf'},
                'cotizacion_voluntaria_afiliado': {'code': 'PensionVoluntaria'},
                'cotizacion_voluntaria_empleador':{'code': 'PensionVoluntariaEmp'},
                'fondo_solidaridad':              {'code': 'FondoSolidaridad'},
                'fondo_subsistencia':             {'code': 'FondoSubsistencia'},
                'valor_no_retenido':              {'code': 'ValorNoRetenido'},
                'total_aportes':                  {'code': 'TotalAportes'},
                'valor_upc':                      {'code': 'Valorupc'},
                'n_autorizacion_incap_eg':        {'code': 'IncapacidadEG'},
                'valor_incapacidad_eg':           {'code': 'ValorIncapacidadEG'},
                'valor_licencia_maternidad':      {'code': 'LicenciaMP'},
                'valor_ccf':                      {'code': 'CotizacionCCF'},
                'ibc_pension':                    {'code': 'IbcPension'},
                'ibc_salud':                      {'code': 'IbcSalud'},
                'ibc_riesgo':                     {'code': 'IbcRiesgo'},
                'ibc_ccf':                        {'code': 'IbcCcf'},
                'ibc_otros_parafiscales':         {'code': 'IbcOtros'},
                'valor_cotizacion_sena':          {'code': 'Sena'},
                'valor_cotizacion_icbf':          {'code': 'CotizacionICBF'},
                'valor_cotizacion_esap':          {'code': 'CotizacionESAP'},
                'valor_cotizacion_men':           {'code': 'CotizacionMEN'},
                'exonerado_1607':                 {'code': 'Exonerado1607'},
            }

            # Inicializar resultados
            #valores_reglas = {k: 0.0 for k in reglas_excel.keys()}
            valores_reglas = {k: "" for k in reglas_excel.keys()}

            # Recorrer líneas de nómina
            for line in payslip.line_ids:
                if not line.salary_rule_id:
                    continue

                rule_code = line.salary_rule_id.code

                for key, config in reglas_excel.items():
                    if rule_code == config['code']:
                        # Si la celda está vacía, asignamos el primer valor
                        if valores_reglas[key] == "":
                            valores_reglas[key] = line.total
                        else:
                            # Si ya tiene un valor numérico, sumamos el siguiente
                            valores_reglas[key] += line.total"""

            claves_reporte = [
                'valor_cotizacion_pension', 'valor_cotizacion_salud', 'valor_cotizacion_riesgo',
                'valor_cotizacion_ccf', 'cotizacion_voluntaria_afiliado', 'cotizacion_voluntaria_empleador',
                'fondo_solidaridad', 'fondo_subsistencia', 'valor_no_retenido', 'total_aportes',
                'valor_upc', 'valor_incapacidad_eg', 'valor_licencia_maternidad', 'ibc_pension',
                'ibc_salud', 'ibc_riesgo', 'ibc_ccf', 'ibc_otros_parafiscales',
                'valor_cotizacion_sena', 'valor_cotizacion_icbf', 'valor_cotizacion_esap',
                'valor_cotizacion_men', 'exonerado_1607'
            ]

            # 2. Inicializar con 0.0 (es mejor para cálculos numéricos)
            valores_reglas = {k: 0.0 for k in claves_reporte}

            # 3. Recorrer las líneas de la nómina
            for line in payslip.line_ids:
                # Obtenemos la marca de la regla
                tipo = line.salary_rule_id.tipo_reporte_excel
                
                # Si la regla tiene una marca y esa marca está en nuestras claves
                if tipo and tipo in valores_reglas:
                    valores_reglas[tipo] += line.total

            ######################Otras tarifas ##################################
            admins = payslip.contract_id.administradoras_ids
            #t_ccf = sum(admins.mapped('tarifa_ccf')) or 0.0
            t_sena = sum(admins.mapped('tarifa_sena')) or 0.0
            t_icbf = sum(admins.mapped('tarifa_icbf')) or 0.0
            t_esap = sum(admins.mapped('tarifa_esap')) or 0.0
            t_men = sum(admins.mapped('tarifa_men')) or 0.0

            ################Clase ################################

            valor_clase = contract.clase if contract.clase else ''

            ################### centro trabajo ################################

            #nombre_departamento = contract.department_id.name if contract.department_id else ''
            val_centro_trabajo = ""
            if contract.work_center_id:
                val_centro_trabajo = contract.work_center_id.name.upper()
            else:
                # SI NO HAY REGISTRO, PASAMOS EL VALOR FIJO
                val_centro_trabajo = "RIESGO III"

            #####################Actividad Economica #######################################
            actividad_economica = contract.economic_activitity if contract.economic_activitity else ''

            ################################upc adicional################################
            mapeo_id = {
                'Cédula de ciudadanía': 'CC',
                'Tarjeta de identidad': 'TI',
                'Registro civil': 'RC',
                'Cédula de extranjería': 'CE',
                'Pasaporte': 'PA',
                'NIT': 'NI',
                'Permiso Especial de Permanencia': 'PE',
                'Permiso por Protección Temporal': 'PT',
            }

            tipo_doc_upc_rec = employee.l10n_latam_identification_type_id

            # 2. Extraemos el nombre (o cadena vacía si no hay)
            nombre_doc_upc = tipo_doc_upc_rec.name or ''

            # 3. Buscamos la abreviatura en el mapeo. Si no está, dejamos el nombre original.
            tipo_doc_upc_abreviado = mapeo_id.get(nombre_doc_upc, nombre_doc_upc)

            # 4. Obtenemos el número de identificación adicional
            numero_upc = employee.upc_identification_number or ''

            data = [
                row_counter,                                      # 1 (A)
                tipo_doc_abreviado or 'CC',                                # 2 (B) - Debe ser 'CC', 'CE', etc.
                str(employee.employee_address_home.vat or '').strip(), # 3 (C) - Sin espacios
                (employee.employee_address_home.last_name or '').upper(), # 4 (D)
                (employee.employee_address_home.second_last_name or '').upper(), # 5 (E)
                (employee.employee_address_home.first_name or '').upper(), # 6 (F)
                (employee.employee_address_home.middle_name or '').upper(), # 7 (G)
                (employee.employee_address_home.state_id.name or 'BOGOTA').upper(), # 8 (H)
                (employee.employee_address_home.city_id.name or 'BOGOTA').upper(), # 9 (I)
                tipo_cotizante_excel_label or '1. DEPENDIENTE',   # 10 (J)
                sub_cotizante_excel_label or 'NINGUNO',           # 11 (K)
                horas_laboradas or 0,                             # 12 (L)
                es_extranjero_label or 'NO',                      # 13 (M)
                es_residente_label or 'NO',                       # 14 (N)
                fecha_radicacion or '',                           # 15 (O)
                
                # --- BLOQUE DE MARCAS (SI/NO) ---
                valor_ing or 'NO',                                # 16 (P) ING
                fecha_inicio_contrato or '',                      # 17 (Q) Fecha ING
                valor_ret or 'NO',                                # 18 (R) RET
                fecha_final_contrato or '',                       # 19 (S) Fecha RET
                novelty_values.get('TDE', 'NO'),                  # 20 (T)
                novelty_values.get('TAE', 'NO'),                  # 21 (U)
                novelty_values.get('TDP', 'NO'),                  # 22 (V)
                novelty_values.get('TAP', 'NO'),                  # 23 (W)
                valor_vsp or 'NO',                                # 24 (X) VSP
                fecha_vsp_str or '',                              # 25 (Y) Fecha VSP
                'NO',                                             # 26 (Z) VST (Variación Transitoria)
                
                # --- BLOQUE AUSENCIAS (3 columnas por cada una) ---
                absence_novelties['SLN']['value'], absence_novelties['SLN']['start'], absence_novelties['SLN']['end'], # 27,28,29
                absence_novelties['IGE']['value'], absence_novelties['IGE']['start'], absence_novelties['IGE']['end'], # 30,31,32
                absence_novelties['LMA']['value'], absence_novelties['LMA']['start'], absence_novelties['LMA']['end'], # 33,34,35
                absence_novelties['VAL-LR']['value'], absence_novelties['VAL-LR']['start'], absence_novelties['VAL-LR']['end'], # 36,37,38
                absence_novelties['AVP']['value'],                # 39 (Solo marca)
                absence_novelties['VCT']['value'], absence_novelties['VCT']['start'], absence_novelties['VCT']['end'], # 40,41,42
                absence_novelties['IRL']['value'], absence_novelties['IRL']['start'], absence_novelties['IRL']['end'], # 43,44,45

                correction_status_excel_value or 'NO',            # 46
                salario_mensual or 0,                             # 47
                integral_excel_value or 'NO',                     # 48
                variable_excel_value or 'NO',                     # 49

                # --- PENSION ---
                pension_admin.get_pension_label() if pension_admin else 'NINGUNA', # 50
                dias_cotizados_pension or 0,                      # 51
                valores_reglas['ibc_pension'] or 0,               # 52
                contract.get_tarifa_by_type('pension') or 0,      # 53
                valores_reglas['valor_cotizacion_pension'] or 0,  # 54
                alto_riesgo_label or 'NO',                        # 55
                valores_reglas['cotizacion_voluntaria_afiliado'] or 0, # 56
                valores_reglas['cotizacion_voluntaria_empleador'] or 0,# 57
                valores_reglas['fondo_solidaridad'] or 0,         # 58
                valores_reglas['fondo_subsistencia'] or 0,        # 59
                valores_reglas['valor_no_retenido'] or 0,         # 60
                valores_reglas['total_aportes'] or 0,             # 61
                pension_admin.get_pension_destino_label() if pension_admin else 'NINGUNA', # 62

                # --- SALUD ---
                salud_admin.get_salud_label() if salud_admin else 'NINGUNA', # 63
                dias_cotizados_salud or 0,                        # 64
                valores_reglas['ibc_salud'] or 0,                 # 65
                contract.get_tarifa_by_type('salud') or 0,        # 66
                valores_reglas['valor_cotizacion_salud'] or 0,    # 67
                valores_reglas['valor_upc'] or 0,                 # 68
                count_ige or '',                                  # 69 N° Autorización EG
                valores_reglas['valor_incapacidad_eg'] or 0,      # 70
                count_lma or '',                                  # 71 N° Autorización LMA
                valores_reglas['valor_licencia_maternidad'] or 0, # 72
                salud_admin.get_salud_destino_label() if salud_admin else 'NINGUNA', # 73

                # --- RIESGOS (ARL) ---
                arl_admin.get_arl_label() if arl_admin else 'NINGUNA', # 74
                dias_cotizados_arl or 0,                          # 75
                valores_reglas['ibc_riesgo'] or 0,                # 76
                contract.get_tarifa_by_type('arl') or 0,          # 77
                valor_clase or '1',                               # 78
                #nombre_departamento or '',                        # 79
                val_centro_trabajo or '',                        # 79
                actividad_economica or '',                        # 80
                valores_reglas['valor_cotizacion_riesgo'] or 0,   # 81

                # --- CAJA Y PARAFISCALES ---
                dias_cotizados_ccf or 0,                          # 82
                ccf_admin.get_ccf_label() if ccf_admin else 'NINGUNA', # 83
                valores_reglas['ibc_ccf'] or 0,                   # 84
                contract.get_tarifa_by_type('ccf') or 0,          # 85
                valores_reglas['valor_cotizacion_ccf'] or 0,      # 86
                valores_reglas['ibc_otros_parafiscales'] or 0,    # 87
                t_sena or 0,                                      # 88
                valores_reglas['valor_cotizacion_sena'] or 0,     # 89
                t_icbf or 0,                                      # 90
                valores_reglas['valor_cotizacion_icbf'] or 0,     # 91
                t_esap or 0,                                      # 92
                valores_reglas['valor_cotizacion_esap'] or 0,     # 93
                t_men or 0,                                       # 94
                valores_reglas['valor_cotizacion_men'] or 0,      # 95
                valores_reglas['exonerado_1607'] or 0,         # 96
                tipo_doc_upc_abreviado or 'CC',                   # 97
                numero_upc or '',                                 # 98
            ]
            
    
            """for col_num, cell_value in enumerate(data):
                # Mostrar las fechas en formato dd/mm/yyyy
                if isinstance(cell_value, (fields.Date, date, datetime)):
                    if isinstance(cell_value, date) and not isinstance(cell_value, datetime):
                        cell_value = datetime.combine(cell_value, datetime.min.time())
                    sheet.write_datetime(row_num, col_num, cell_value, date_format)
                else:
                    sheet.write(row_num, col_num, cell_value)"""

            # --- BUCLE DE ESCRITURA AJUSTADO ---
            """for col_num, cell_value in enumerate(data):
                # Mostrar las fechas en formato dd/mm/yyyy
                if isinstance(cell_value, (fields.Date, date, datetime)):
                    if isinstance(cell_value, date) and not isinstance(cell_value, datetime):
                        cell_value = datetime.combine(cell_value, datetime.min.time())
                    sheet.write_datetime(row_num, col_num, cell_value, date_format)
                else:
                    sheet.write(row_num, col_num, cell_value)"""

            # --- PROCESO DE ESCRITURA CON VALIDACIÓN DE FORMATO ---
            for col_num, cell_value in enumerate(data, start=1):
                cell = sheet.cell(row=row_num, column=col_num)

                # 1. COLUMNA DE ÍNDICE (Registro)
                if col_num == 1:
                    # Usamos row_counter para poner 1, 2, 3... correlativamente
                    cell.value = row_counter
                    cell.number_format = '0' # Formato General sin decimales
                    # Alineación a la derecha para que se vea ordenado
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    continue # Pasamos a la siguiente columna



                # --- NUEVO BLOQUE: COLUMNAS CON SIGNO $ Y ALINEACIÓN DERECHA ---
                # Agregamos todas las que pediste: 47, 52, 54, 56-61, 65, 68
                elif col_num in [47, 52, 54, 56, 57, 58, 59, 60, 61, 65, 68]:
                    try:
                        num_val = float(cell_value or 0)
                        if num_val != 0:
                            cell.value = int(num_val)
                            # Formato con signo $, miles con punto, sin decimales
                            cell.number_format = '"$"#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.value = "" # Celda limpia si es cero
                    except:
                        cell.value = ""
                    continue # Saltamos a la siguiente columna

                # --- COLUMNA 96: TRANSFORMACIÓN DE NÚMERO A SI/NO ---
                elif col_num == 96:
                    try:
                        # Evaluamos el número que viene (ej. exonerado_1607)
                        num_val = float(cell_value or 0)
                        
                        if num_val > 0:
                            cell.value = "SI"
                        else:
                            cell.value = "NO"
                    except:
                        # Respaldo por si viene texto
                        val_str = str(cell_value or '').strip().upper()
                        cell.value = "SI" if val_str in ['SI', 'TRUE', 'S', '1'] else "NO"
                    
                    continue

                if col_num in [50, 62,63,73, 74, 83]:
                    val_admin = str(cell_value or '').strip().upper()
                    if val_admin in ['', 'NONE', 'FALSE', '0', 'N/A', 'NINGUNO']:
                        cell.value = "NINGUNA"
                    else:
                        cell.value = val_admin

                # --- 1. CLASE DE RIESGO (Columna 78) ---
                elif col_num == 55:
                    # Si el valor es '5', está vacío o es falso -> "Sin Riesgo"
                    val_riesgo = str(cell_value or '').strip()
                    if val_riesgo in ['', '5', 'NONE', 'FALSE', '0', 'Sin riesgo']:
                        cell.value = "Sin Riesgo"
                    else:
                        # Si por alguna razón viene el texto de la selección de Odoo
                        cell.value = val_riesgo
                    continue # Saltamos a la siguiente columna


                elif col_num == 68:
                    try:
                        val_upc = float(cell_value or 0)
                        # Si hay valor, ponemos el signo $. Si es 0, lo dejamos con $0 o limpio.
                        cell.value = int(val_upc)
                        cell.number_format = '"$"#,##' 
                    except:
                        cell.value = 0
                        cell.number_format = '"$"#,##'  
                    continue # Saltamos a la siguiente columna     

                # --- 2. SALARIO INTEGRAL (48) Y VARIABLE (49) ---
                elif col_num in [48, 49]:
                    val_bool = str(cell_value or '').strip().upper()
                    if val_bool in ['SI', 'TRUE', '1', 'S']:
                        cell.value = "SI"
                    else:
                        cell.value = "NO"
                    continue # Saltamos a la siguiente columna



                # --- 2. TARIFAS (Formato % con 3 decimales) ---
                elif col_num in [53, 66, 77, 85, 88, 90, 92, 94]:
                    try:
                        clean_val = str(cell_value or 0).replace('%', '').strip()
                        val = float(clean_val)
                        cell.value = val / 100 if val > 1 else val
                        cell.number_format = '0.00%'
                    except:
                            cell.value = 0
                            cell.number_format = '0.00%'
                    continue # Saltamos a la siguiente columna


                # --- 3. NOMBRES Y APELLIDOS (4, 5, 6, 7) ---
                # Si no hay valor, queda VACÍO, no pone "NO"
                elif col_num in [4, 5, 6, 7]:
                    val_name = str(cell_value or '').strip().upper()
                    cell.value = "" if val_name in ['', 'NONE', 'FALSE'] else val_name
                    continue # Saltamos a la siguiente columna


                # --- 4. IDENTIFICACIONES Y UPC ADICIONAL (Col: 3, 97, 98) ---
                elif col_num in [2,3, 97, 98]:
                    val_id = str(cell_value or '').strip().upper()
                    # En la 98 NUNCA va "NINGUNO", queda vacío
                    cell.value = "" if val_id in ['', 'NONE', 'FALSE', 'NINGUNO'] else val_id
                    cell.data_type = 's' # Forzar texto
                    continue # Saltamos a la siguiente columna

                if col_num == 79:
                    val_centro = str(cell_value or '').strip().upper()
                    
                    # Si no hay valor, asignamos el valor fijo "RIESGO III"
                    if val_centro in ['', 'NONE', 'FALSE', '0', '0.0', 'NO']:
                        cell.value = "RIESGO III"
                    else:
                        cell.value = val_centro
                    
                    # 1. FORZAR FORMATO TEXTO: Esto le dice a Excel que trate el contenido como String
                    cell.data_type = 's'
                    
                    # 2. ALINEACIÓN: Izquierda para que se lea correctamente el nombre
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 3. SALTO CRÍTICO: El 'continue' evita que el bloque 'else' le ponga un "NO"
                    # y evita que el bloque de números intente convertirlo a float.
                    continue

                # --- COLUMNA 80: ACTIVIDAD ECONÓMICA ---
                elif col_num == 80:
                    val_actividad = str(cell_value or '').strip()
                    
                    # Si no hay valor, la celda queda VACÍA (blanca)
                    if val_actividad.upper() in ['', 'NONE', 'FALSE', '0', '0.0']:
                        cell.value = ""
                    else:
                        cell.value = val_actividad
                    
                    # Forzamos formato texto para que no se pierdan ceros iniciales
                    cell.data_type = 's'
                    
                    # IMPORTANTE: El continue evita que pase al "else" final y le ponga "NO"
                    continue

                # --- 5. DÍAS Y VALORES MONETARIOS (Limpios si son 0) ---
                elif (47 <= col_num <= 78) or (81 <= col_num <= 95) or col_num == 12:
                    try:
                        num_val = float(cell_value or 0)
                        if num_val != 0:
                            cell.value = int(num_val)
                            cell.number_format = '#,##0'
                            
                            # ALINEACIÓN A LA DERECHA SOLO PARA LA COLUMNA 12
                            if col_num == 12:
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            else:
                                # Las demás (47 a 96) mantienen la alineación estándar (izquierda o general)
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.value = "" 
                    except:
                        cell.value = ""
                    continue # <--- ESTO EVITA QUE PASE AL 'ELSE' DE ABAJO

                # --- 6. FECHAS (Regla 3 Reforzada) ---
                elif isinstance(cell_value, (date, datetime)):
                    cell.value = cell_value
                    cell.number_format = 'yyyy-mm-dd'

                # --- 7. CASO GENERAL (Marcas SI/NO y Fechas en String) ---
                else:
                    val_gen = str(cell_value or '').strip()
                    if val_gen.upper() in ['NONE', 'FALSE', '']:
                        # Si es una columna de fecha (ej. 17, 19, 25) y está vacía, dejar limpia
                        if col_num in [15,17, 19, 25, 28, 29, 31, 32, 34, 35,37,38,41,42,44,45]: 
                            cell.value = ""
                        else:
                            cell.value = "NO"
                    else:
                        cell.value = val_gen.upper()

                    # --- AJUSTE DE ALINEACIÓN ESPECÍFICO ---
                    if col_num in [25, 34, 35]:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            row_num += 1
            row_counter += 1

        # 5. AJUSTE DE ANCHO Y GUARDADO FINAL
        for col in sheet.columns:
            column_letter = col[0].column_letter
            sheet.column_dimensions[column_letter].width = 18

        wb.save(output)
        output.seek(0)

        # --- 2. NOMBRE FIJO DEL ARCHIVO ---
        # Forzamos el nombre exacto que me indicas
        file_name = "DIVITIASSAS_1.xlsx"


        # --- 3. CREAR EL ADJUNTO ---
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()), 
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # --- 4. CERRAR EL FLUJO ---
        output.close()

        # --- 5. DEVOLVER LA ACCIÓN DE DESCARGA ---
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % (attachment.id),
            'target': 'self',
        }


   